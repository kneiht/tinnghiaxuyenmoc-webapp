import pandas as pd
from slugify import slugify
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.apps import apps
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .forms import *
from .models.models import *
from .renders import *
from .utils import *

from django.utils import timezone
import pytz


@login_required
def download_excel(request, model_name):
    """
    Download an Excel file with data from the specified model using pandas.
    The headers are the field verbose names, or field names if no verbose name exists.
    Uses the model's excel_columns attribute if available.
    """
    if request.method != "POST":
        return HttpResponse("This view only accepts POST requests", status=405)

    try:
        # Get the model class based on the model_name
        model_class = apps.get_model("app_dashboard", model_name)

        # Get the primary key field name
        pk_field = model_class._meta.pk.name
        # Get all fields including primary key
        fields = [pk_field] + [field.name for field in model_class._meta.fields]

        # Remove fields if they exist
        if "last_saved" in fields:
            fields.remove("last_saved")
        if "archived" in fields:
            fields.remove("archived")
        if "secondary_id" in fields:
            fields.remove("secondary_id")

        # Get verbose names for headers
        headers = []
        for field_name in fields:
            field = model_class._meta.get_field(field_name)
            headers.append(getattr(field, "verbose_name", field_name))

        # Get project_id and records
        project_id = request.POST.get("project_id", None)

        try:
            project_id = int(project_id)
            if project_id < 1:
                project_id = None
        except:
            project_id = None

        # Get start_date and end_date from POST data
        start_date_str = request.POST.get("start_date", None)
        end_date_str = request.POST.get("end_date", None)

        # Validate and parse dates
        try:
            start_date = get_valid_date(start_date_str, "none")
            end_date = get_valid_date(end_date_str, "none")
        except ValueError:
            start_date = None
            end_date = None

        if project_id:
            project = Project.objects.get(id=project_id)
            records = model_class.objects.filter(project=project)
        else:
            records = model_class.objects.all()

        print("project_id", project_id)
        print("records", records.count())
        print("start_date", start_date)
        print("end_date", end_date)

        # Apply date filtering if valid dates are provided
        if start_date and end_date:
            records = records.filter(created_at__date__range=[start_date, end_date])
        elif start_date:
            records = records.filter(created_at__date__gte=start_date)
        elif end_date:
            records = records.filter(created_at__date__lte=end_date)

        records = records.values(*fields).order_by(pk_field)

        # order by id
        records = records.order_by(pk_field)

        # Create a DataFrame from the records
        df = pd.DataFrame(list(records))

        # If DataFrame is empty, create one with just the headers
        if df.empty and fields:
            df = pd.DataFrame(columns=fields)

        # Convert choice fields to their display values and handle foreign keys
        for field_name in fields:
            field = model_class._meta.get_field(field_name)
            if hasattr(field, "choices") and field.choices:
                choices_dict = dict(field.choices)
                df[field_name] = df[field_name].map(lambda x: choices_dict.get(x, x))
            elif field.is_relation and field.many_to_one:
                # This checks if it's a ForeignKey
                # Get the related model
                related_model = field.related_model
                # Create a dictionary of id to string representation
                related_objects = related_model.objects.in_bulk()
                df[field_name] = df[field_name].map(
                    lambda x: (
                        f"ID={x}:{related_objects[x]}" if x in related_objects else x
                    )
                )

        # Rename columns to use verbose names
        column_mapping = {fields[i]: headers[i] for i in range(len(fields))}
        df = df.rename(columns=column_mapping)

        # if hash vietnamese_name exists, use it as the file name
        if hasattr(model_class, "vietnamese_name"):
            filename = model_class.vietnamese_name
            # slugging the filename
            filename = slugify(filename)
        else:
            filename = model_name

        # Add datetime to filename
        vietnam_tz = pytz.timezone("Asia/Ho_Chi_Minh")
        current_time = timezone.now().astimezone(vietnam_tz).strftime("%Y%m%d_%H%M%S")
        filename = f"{filename}_{current_time}"

        # Create a response with Excel file
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'

        # Write the DataFrame to Excel
        with pd.ExcelWriter(response, engine="openpyxl") as writer:
            # Write data sheet
            df.to_excel(writer, index=False, sheet_name="Data")

            # Create instruction sheet
            instructions = pd.DataFrame(
                [
                    ["Hướng dẫn sử dụng:"],
                    ["1. Không thay đổi tên các cột"],
                    ["2. Không xóa cột ID (nếu có)"],
                    ["3. Để thêm mới dữ liệu: để trống cột ID"],
                    ["4. Để cập nhật dữ liệu: giữ nguyên ID"],
                    ["5. Với các trường liên kết (dạng ID=123:Tên), phải nhập đúng ID"],
                    ["6. Các trường bắt buộc không được để trống"],
                    ["7. Định dạng ngày tháng: YYYY-MM-DD"],
                    [
                        "8. File Excel này được tạo lúc: "
                        + current_time
                        + "Tải lại file nếu có dữ liệu mới hơn trên server"
                    ],
                ],
                columns=["Instructions"],
            )
            instructions.to_excel(writer, index=False, sheet_name="Instructions")

            # Add extra sheet for SupplyOrderSupply if model_name is SupplyOrder
            if model_name == "SupplyOrder":
                # Import the SupplyOrderSupply model
                from .models.supply import SupplyOrderSupply

                # Get all SupplyOrderSupply records related to the SupplyOrder records
                supply_order_ids = list(records.values_list("id", flat=True))

                if supply_order_ids:
                    # Get all related SupplyOrderSupply records
                    supply_items = SupplyOrderSupply.objects.filter(
                        supply_order_id__in=supply_order_ids
                    )

                    # Get fields for SupplyOrderSupply
                    supply_fields = [
                        field.name
                        for field in SupplyOrderSupply._meta.fields
                        if field.name not in ["last_saved", "archived", "secondary_id"]
                    ]

                    # Get verbose names for headers
                    supply_headers = []
                    for field_name in supply_fields:
                        field = SupplyOrderSupply._meta.get_field(field_name)
                        supply_headers.append(
                            getattr(field, "verbose_name", field_name)
                        )

                    # Get the data
                    supply_records = supply_items.values(*supply_fields)

                    # Create DataFrame
                    supply_df = pd.DataFrame(list(supply_records))

                    # If DataFrame is empty, create one with just the headers
                    if supply_df.empty and supply_fields:
                        supply_df = pd.DataFrame(columns=supply_fields)

                    # Convert choice fields to their display values and handle foreign keys
                    for field_name in supply_fields:
                        field = SupplyOrderSupply._meta.get_field(field_name)
                        if hasattr(field, "choices") and field.choices:
                            choices_dict = dict(field.choices)
                            supply_df[field_name] = supply_df[field_name].map(
                                lambda x: choices_dict.get(x, x)
                            )
                        elif field.is_relation and field.many_to_one:
                            # This checks if it's a ForeignKey
                            # Get the related model
                            related_model = field.related_model
                            # Create a dictionary of id to string representation
                            related_objects = related_model.objects.in_bulk()
                            supply_df[field_name] = supply_df[field_name].map(
                                lambda x: (
                                    f"ID={x}:{related_objects[x]}"
                                    if x in related_objects
                                    else x
                                )
                            )

                    # Rename columns to use verbose names
                    supply_column_mapping = {
                        supply_fields[i]: supply_headers[i]
                        for i in range(len(supply_fields))
                    }
                    supply_df = supply_df.rename(columns=supply_column_mapping)

                    # Write to Excel
                    supply_df.to_excel(writer, index=False, sheet_name="Extra")

                    # Style extra sheet
                    worksheet_extra = writer.sheets["Extra"]

                    # Format header row
                    for col_num, column in enumerate(supply_df.columns, 1):
                        cell = worksheet_extra.cell(row=1, column=col_num)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(
                            start_color="FFFFD4", end_color="FFFFD4", fill_type="solid"
                        )

                    # Adjust column widths
                    for col_num, column in enumerate(supply_df.columns, 1):
                        column_letter = get_column_letter(col_num)
                        worksheet_extra.column_dimensions[column_letter].width = 15

            # Add extra sheet for SubJobOrderSubJob if model_name is SubJobOrder
            elif model_name == "SubJobOrder":
                # Import the SubJobOrderSubJob model
                from .models.subcontractor import SubJobOrderSubJob

                # Get all SubJobOrder IDs from the records
                sub_job_order_ids = list(records.values_list("id", flat=True))

                if sub_job_order_ids:
                    # Get all related SubJobOrderSubJob records
                    sub_job_items = SubJobOrderSubJob.objects.filter(
                        sub_job_order_id__in=sub_job_order_ids
                    )

                    # Get fields for SubJobOrderSubJob
                    sub_job_fields = [
                        field.name
                        for field in SubJobOrderSubJob._meta.fields
                        if field.name not in ["last_saved", "archived", "secondary_id"]
                    ]

                    # Get verbose names for headers
                    sub_job_headers = []
                    for field_name in sub_job_fields:
                        field = SubJobOrderSubJob._meta.get_field(field_name)
                        sub_job_headers.append(
                            getattr(field, "verbose_name", field_name)
                        )

                    # Get the data
                    sub_job_records = sub_job_items.values(*sub_job_fields)

                    # Create DataFrame
                    sub_job_df = pd.DataFrame(list(sub_job_records))

                    # If DataFrame is empty, create one with just the headers
                    if sub_job_df.empty and sub_job_fields:
                        sub_job_df = pd.DataFrame(columns=sub_job_fields)

                    # Convert choice fields to their display values and handle foreign keys
                    for field_name in sub_job_fields:
                        field = SubJobOrderSubJob._meta.get_field(field_name)
                        if hasattr(field, "choices") and field.choices:
                            choices_dict = dict(field.choices)
                            sub_job_df[field_name] = sub_job_df[field_name].map(
                                lambda x: choices_dict.get(x, x)
                            )
                        elif field.is_relation and field.many_to_one:
                            # This checks if it's a ForeignKey
                            # Get the related model
                            related_model = field.related_model
                            # Create a dictionary of id to string representation
                            related_objects = related_model.objects.in_bulk()
                            sub_job_df[field_name] = sub_job_df[field_name].map(
                                lambda x: (
                                    f"ID={x}:{related_objects[x]}"
                                    if x in related_objects
                                    else x
                                )
                            )

                    # Rename columns to use verbose names
                    sub_job_column_mapping = {
                        sub_job_fields[i]: sub_job_headers[i]
                        for i in range(len(sub_job_fields))
                    }
                    sub_job_df = sub_job_df.rename(columns=sub_job_column_mapping)

                    # Write to Excel
                    sub_job_df.to_excel(writer, index=False, sheet_name="Extra")

                    # Style extra sheet
                    worksheet_extra = writer.sheets["Extra"]

                    # Format header row
                    for col_num, column in enumerate(sub_job_df.columns, 1):
                        cell = worksheet_extra.cell(row=1, column=col_num)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(
                            start_color="FFFFD4", end_color="FFFFD4", fill_type="solid"
                        )

                    # Adjust column widths
                    for col_num, column in enumerate(sub_job_df.columns, 1):
                        column_letter = get_column_letter(col_num)
                        worksheet_extra.column_dimensions[column_letter].width = 15

            # Add extra sheet for VehicleMaintenanceRepairPart if model_name is VehicleMaintenance
            elif model_name == "VehicleMaintenance":
                # Import the VehicleMaintenanceRepairPart model
                from .models.maintenance import VehicleMaintenanceRepairPart

                # Get all VehicleMaintenance IDs from the records
                vehicle_maintenance_ids = list(records.values_list("id", flat=True))

                if vehicle_maintenance_ids:
                    # Get all related VehicleMaintenanceRepairPart records
                    repair_part_items = VehicleMaintenanceRepairPart.objects.filter(
                        vehicle_maintenance_id__in=vehicle_maintenance_ids
                    )

                    # Get fields for VehicleMaintenanceRepairPart
                    repair_part_fields = [
                        field.name
                        for field in VehicleMaintenanceRepairPart._meta.fields
                        if field.name not in ["last_saved", "archived", "secondary_id"]
                    ]

                    # Get verbose names for headers
                    repair_part_headers = []
                    for field_name in repair_part_fields:
                        field = VehicleMaintenanceRepairPart._meta.get_field(field_name)
                        repair_part_headers.append(
                            getattr(field, "verbose_name", field_name)
                        )

                    # Get the data
                    repair_part_records = repair_part_items.values(*repair_part_fields)

                    # Create DataFrame
                    repair_part_df = pd.DataFrame(list(repair_part_records))

                    # If DataFrame is empty, create one with just the headers
                    if repair_part_df.empty and repair_part_fields:
                        repair_part_df = pd.DataFrame(columns=repair_part_fields)

                    # Convert choice fields to their display values and handle foreign keys
                    for field_name in repair_part_fields:
                        field = VehicleMaintenanceRepairPart._meta.get_field(field_name)
                        if hasattr(field, "choices") and field.choices:
                            choices_dict = dict(field.choices)
                            repair_part_df[field_name] = repair_part_df[field_name].map(
                                lambda x: choices_dict.get(x, x)
                            )
                        elif field.is_relation and field.many_to_one:
                            # This checks if it's a ForeignKey
                            # Get the related model
                            related_model = field.related_model
                            # Create a dictionary of id to string representation
                            related_objects = related_model.objects.in_bulk()
                            repair_part_df[field_name] = repair_part_df[field_name].map(
                                lambda x: (
                                    f"ID={x}:{related_objects[x]}"
                                    if x in related_objects
                                    else x
                                )
                            )

                    # Rename columns to use verbose names
                    repair_part_column_mapping = {
                        repair_part_fields[i]: repair_part_headers[i]
                        for i in range(len(repair_part_fields))
                    }
                    repair_part_df = repair_part_df.rename(
                        columns=repair_part_column_mapping
                    )

                    # Write to Excel
                    repair_part_df.to_excel(writer, index=False, sheet_name="Extra")

                    # Style extra sheet
                    worksheet_extra = writer.sheets["Extra"]

                    # Format header row
                    for col_num, column in enumerate(repair_part_df.columns, 1):
                        cell = worksheet_extra.cell(row=1, column=col_num)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(
                            start_color="FFFFD4", end_color="FFFFD4", fill_type="solid"
                        )

                    # Adjust column widths
                    for col_num, column in enumerate(repair_part_df.columns, 1):
                        column_letter = get_column_letter(col_num)
                        worksheet_extra.column_dimensions[column_letter].width = 15

            # Style data sheet
            workbook = writer.book
            worksheet = writer.sheets["Data"]

            # Format header row
            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="FFFFD4", end_color="FFFFD4", fill_type="solid"
                )

            # Adjust column widths
            for col_num, column in enumerate(df.columns, 1):
                column_letter = get_column_letter(col_num)
                worksheet.column_dimensions[column_letter].width = 15

            # Style instruction sheet
            worksheet_instructions = writer.sheets["Instructions"]
            worksheet_instructions.column_dimensions["A"].width = 100
            for row in range(1, 10):
                cell = worksheet_instructions.cell(row=row, column=1)
                cell.font = Font(size=12)
                if row == 1:
                    cell.font = Font(bold=True, size=14)

        return response

    except Exception as e:
        raise e


@login_required
def upload_excel(request, model_name):

    if not request.user.is_superuser:
        return HttpResponse(
            render_message(
                request,
                "Chỉ có admin mới có quyền tải lên file excel",
                message_type="red",
            )
        )

    if request.method != "POST":
        return HttpResponse("This view only accepts POST requests", status=405)

    try:
        # Get the model class and read file
        model_class = apps.get_model("app_dashboard", model_name)
        excel_file = request.FILES.get("file")
        if not excel_file:
            return HttpResponse(
                render_message(request, "Không phải file excel", message_type="red")
            )

        # Get project_id from POST
        project_id = request.POST.get("project_id", None)

        if project_id:
            try:
                project_id = int(project_id)
                if project_id < 1:
                    project_id = None
            except Exception as e:
                raise e
                project_id = None

        # Check if model has project field and project_id is provided
        if project_id and not hasattr(model_class, "project"):
            return HttpResponse(
                render_message(
                    request, "Model không có trường project", message_type="red"
                )
            )

        try:
            df = pd.read_excel(excel_file, sheet_name="Data")
        except:
            return HttpResponse(
                render_message(
                    request, "File excel không có sheet 'Data'", message_type="red"
                )
            )

        if df.empty:
            return HttpResponse(
                render_message(
                    request, "Sheet 'Data' không có dữ liệu", message_type="red"
                )
            )

        # Setup field mappings
        fields = [
            field.name
            for field in model_class._meta.fields
            if field.name not in ["last_saved", "archived"]
        ]
        field_mapping = {}
        for field_name in fields:
            field = model_class._meta.get_field(field_name)
            verbose_name = getattr(field, "verbose_name", field_name)
            field_mapping[verbose_name] = field_name

        # Check for undefined columns
        undefined_columns = [col for col in df.columns if col not in field_mapping]
        if undefined_columns:
            return HttpResponse(
                render_message(
                    request,
                    f"Không xác định các cột: {', '.join(undefined_columns)}",
                    message_type="red",
                )
            )

        # Create reverse mapping for choice fields
        choice_mappings = {}
        for field_name in fields:
            field = model_class._meta.get_field(field_name)
            if hasattr(field, "choices") and field.choices:
                reverse_choices = {display: value for value, display in field.choices}
                choice_mappings[field_name] = reverse_choices

        df_renamed = df.rename(columns=field_mapping)

        # Convert choice fields back to their internal values and handle foreign keys
        for field_name in fields:
            field = model_class._meta.get_field(field_name)
            if hasattr(field, "choices") and field.choices:
                if field_name in df_renamed.columns:
                    df_renamed[field_name] = df_renamed[field_name].map(
                        lambda x: choice_mappings[field_name].get(x, x)
                    )
            elif (
                field.is_relation
                and (field.many_to_one or field.one_to_one)
                and field_name in df_renamed.columns
            ):
                # Extract ID from the pattern "ID=123:Name" or direct ID
                def get_foreign_key_instance(x):
                    if pd.isna(x):
                        return None
                    try:
                        # Try to extract ID from "ID=123:Name" format
                        if str(x).startswith("ID="):
                            id_value = int(str(x).split("ID=")[1].split(":")[0])
                        else:
                            # If direct ID provided
                            id_value = int(x)
                        # Get the actual instance
                        return field.related_model.objects.get(pk=id_value)
                    except (ValueError, field.related_model.DoesNotExist):
                        return x

                df_renamed[field_name] = df_renamed[field_name].map(
                    get_foreign_key_instance
                )

        errors = []
        instances_to_create = []
        instances_to_update = []
        pk_field = model_class._meta.pk.name

        # First pass: validate all records and collect errors
        for index, row in df_renamed.iterrows():
            try:
                data = {k: v for k, v in row.to_dict().items() if pd.notna(v)}
                record_id = data.get(pk_field)
                # Add project to data if model has project field
                if project_id:
                    data["project_id"] = project_id

                if pd.isna(record_id):
                    # Validate new record
                    instance = model_class(**data)
                    instance.full_clean()
                    instances_to_create.append(instance)
                else:
                    # Validate existing record
                    try:
                        # If project_id exists, only allow updates to records of that project
                        if project_id:
                            instance = model_class.objects.get(
                                pk=record_id, project_id=project_id
                            )
                        else:
                            instance = model_class.objects.get(pk=record_id)

                        for field, value in data.items():
                            setattr(instance, field, value)
                        instance.full_clean()
                        instances_to_update.append(instance)
                    except model_class.DoesNotExist:
                        if project_id:
                            errors.append(
                                f"Dòng {index + 2}: Không tìm thấy dữ liệu với ID = {int(record_id)} trong dự án này"
                            )
                        else:
                            errors.append(
                                f"Dòng {index + 2}: Không tìm thấy dữ liệu với ID = {int(record_id)}"
                            )
            except Exception as e:
                errors.append(f"Dòng {index + 2}: lỗi {str(e)}")

        # If there are any errors, return them without making any changes
        if errors:
            error_message = "\n".join(errors)
            return HttpResponse(
                render_message(request, error_message, message_type="red")
            )

        # If no errors, save all changes in a transaction
        with transaction.atomic():
            for instance in instances_to_create:
                instance.save()
            for instance in instances_to_update:
                instance.save()

        message = f"Tải lên dữ liệu excel thành công. \nTạo: {len(instances_to_create)}, \nCập nhật: {len(instances_to_update)}"
        html_message = render_message(
            request, message, message_type="green", ok_button_function="reload"
        )
        return HttpResponse(html_message)

    except Exception as e:
        message = f"Error: {str(e)}"
        return HttpResponse(render_message(request, message, message_type="red"))

