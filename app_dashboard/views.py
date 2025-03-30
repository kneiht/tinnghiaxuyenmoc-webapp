import os
import json
import tempfile
import shutil
import requests
import pandas as pd
import traceback

from requests.auth import HTTPBasicAuth
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, render, redirect

from core import settings

from .forms import *
from .models.models import *
from .renders import *
from .utils import *


@login_required
def decide_permission(request, action, params):
    # CHECK PERMISSIONS
    tab = params.get("tab", None)
    if not tab:
        model = params.get("model", None)
        sub_page = model
    elif tab == "vehicle_operation_data_by_date":
        sub_page = "VehicleOperationRecord"
    elif tab == "driver_salary":
        sub_page = "ConstructionDriverSalary"
    elif tab == "vehicle_revenue":
        sub_page = "ConstructionReportPL"

    # FEATURES THAT ARE DEVELOPING
    if sub_page in ("Task", "Announcement"):
        message = "Không tìm thấy chức năng này. \n Có thể chức năng này đang được phát triển!"
        return render(request, "components/message_page.html", {"message": message})

    user = request.user
    permission = user.check_permission(sub_page)
    if action == "read":
        if not permission.read:
            message = "Bạn chưa được cấp quyền truy cập trang này. \n Vui lòng liên hệ admin cấp quyền."
            return render(request, "components/message_page.html", {"message": message})
        else:
            return None

    elif action == "create":
        if not permission.create:
            message = "Tạo dữ liệu không thành công vì chưa được cấp quyền. \n\n Vui lòng liên hệ admin cấp quyền."
            message_type = "red"
            return render_message(request, message=message, message_type=message_type)
        else:
            if model == "SupplyOrder":
                # Check if project user
                project_id = get_valid_id(params.get("project_id", 0))
                project = Project.objects.filter(pk=project_id).first()
                if not project:
                    message = "Dự án không tồn tại. \n Vui lòng liên hệ admin."
                    message_type = "red"
                    return render_message(
                        request, message=message, message_type=message_type
                    )

                project_user = ProjectUser.objects.filter(
                    user=user, project=project
                ).first()
                if not project_user or project_user.role not in [
                    "supervisor",
                    "accountant",
                ]:
                    message = "Chỉ có vị trí Giám sát và Kế toán mới có thể tạo và sửa phiếu. \n Vui lòng liên hệ admin cấp quyền."
                    message_type = "red"
                    return render_message(
                        request, message=message, message_type=message_type
                    )

            return None

    elif action == "update":
        if not permission.update:
            message = "Cập nhật dữ liệu không thành công vì chưa được cấp quyền. \n\n Vui lòng liên hệ admin cấp quyền."
            message_type = "red"
            return render_message(request, message=message, message_type=message_type)
        else:
            if model == "SupplyOrder":
                # Check if project user
                project_id = get_valid_id(params.get("project_id", 0))
                project = Project.objects.filter(pk=project_id).first()
                if not project:
                    message = "Dự án không tồn tại. \n Vui lòng liên hệ admin."
                    message_type = "red"
                    return render_message(
                        request, message=message, message_type=message_type
                    )

                project_user = ProjectUser.objects.filter(
                    user=request.user, project=project
                ).first()
                if not project_user or project_user.role not in [
                    "supervisor",
                    "accountant",
                ]:
                    message = "Chỉ có vị trí Giám sát và Kế toán mới có thể tạo và sửa phiếu. \n Vui lòng liên hệ admin cấp quyền."
                    message_type = "red"
                    return render_message(
                        request, message=message, message_type=message_type
                    )

            return None

    elif action == "delete":
        if not permission.delete:
            message = "Không thể xóa dữ liệu vì chưa được cấp quyền. \n\n Vui lòng liên hệ admin cấp quyền."
            message_type = "red"
            return render_message(request, message=message, message_type=message_type)
        else:
            return None

    elif action == "approve":
        if not permission.delete:
            message = "Không thể thay đổi trạng thái duyệt, vì chưa được cấp quyền. \n\n Vui lòng liên hệ admin cấp quyền."
            message_type = "red"
            return render_message(request, message=message, message_type=message_type)
        else:
            return None
    else:
        return None


# HANDLE FORMS ===============================================================
@login_required
def handle_form(request, model, pk=0):
    print(request.POST)
    # Todo: should have list of model that can be accessed
    # Convert model name to model class
    model_class = globals()[model]
    form_class = globals()[model + "Form"]

    # check if not Post => return 404
    if request.method != "POST":
        return HttpResponseForbidden()

    # project_id
    project_id = get_valid_id(request.POST.get("project", 0))

    # Get form
    instance = model_class.objects.filter(pk=pk).first()
    form = form_class(request.POST, request.FILES, instance=instance)

    # Add missing data if there's an instance
    # Add missing data if there's an instance
    if instance:
        post_data = request.POST.copy()
        for field_name in form_class.Meta.fields:
            if not post_data.get(field_name) and hasattr(instance, field_name):
                field_value = getattr(instance, field_name)
                if field_value:
                    # Check if field_value is a model instance and convert to ID
                    if hasattr(field_value, 'pk'):
                        post_data[field_name] = field_value.pk
                    else:
                        post_data[field_name] = field_value
        form = form_class(post_data, request.FILES, instance=instance)

    # Delete and restore
    if request.POST.get("archived") == "true":
        # CHECK PERMISSIONS
        forbit_html = decide_permission(request, "delete", {"model": model})
        if forbit_html:
            return HttpResponse(forbit_html)

        try:
            # if record is VechicleMaintenance":
            if (
                model != "VehicleMaintenance"
                and model != "SupplyOrder"
                and model != "SubJobOrder"
            ):
                record = instance
                # Get related records
                related_records = []
                for related_object in record._meta.related_objects:
                    related_manager = getattr(
                        record, related_object.get_accessor_name()
                    )
                    related_records.extend(list(related_manager.all()))

                # If there are related records, show warning
                if related_records:
                    # Create list of related records
                    related_records_info = []
                    for related_record in related_records:
                        try:
                            display_name = related_record.vietnamese_name
                        except:
                            display_name = related_record._meta.verbose_name
                        record_info = f"{display_name}: {str(related_record)}"
                        related_records_info.append(record_info)

                    # Create message with related records
                    message = "Không thể xóa dữ liệu vì có các bản ghi liên quan.\n Nếu muốn xóa dữ liệu này, phải xóa các dữ liệu liên quan trước để đảm bảo toàn vẹn dữ liệu:\n\n"
                    message += "🔹 " + "\n🔹 ".join(related_records_info)

                    html_message = render_message(
                        request,
                        message=message,
                        message_type="red",
                    )
                    return HttpResponse(html_message)

            record = instance
            record.style = "hidden"
            html_message = render_message(
                request,
                message="Xóa dữ liệu thành công.",
                message_type="green",
            )
            html_record = render_display_records(
                request,
                model=model,
                records=[record],
                update="True",
                project_id=project_id,
            )
            record.delete()
            return HttpResponse(html_message + html_record)
        except Exception as e:
            html_message = render_message(
                request,
                message="Xóa dữ liệu thất bại.\n\n" + str(e),
                message_type="red",
            )
            # panic
            print(traceback.format_exc())

            return HttpResponse(html_message)
    elif request.POST.get("archived") == "false":
        # CHECK PERMISSIONS
        forbit_html = decide_permission(request, "update", {"model": model})
        if forbit_html:
            return HttpResponse(forbit_html)
        html_message = render_message(
            request,
            message="Khôi phục dữ liệu thất bại.\n\nChức năng này đang thử nghiệm.",
            message_type="red",
        )
        return HttpResponse(html_message + html_record)

    if instance is None:  # This is a new form
        # Handle the case of the project is created and need to be assigned to a user
        # instance_form.user = request.user
        # CHECK PERMISSIONS
        forbit_html = decide_permission(
            request, "create", {"model": model, "project_id": project_id}
        )
        if forbit_html:
            return HttpResponse(forbit_html)

        if (
            model == "VehicleMaintenance"
            or model == "SupplyOrder"
            or model == "SubJobOrder"
        ):
            # set approval status to make sure there no injection hacking
            form.instance.approval_status = "scratch"

    else:  # update
        # CHECK PERMISSIONS
        forbit_html = decide_permission(
            request, "update", {"model": model, "project_id": project_id}
        )
        if forbit_html:
            return HttpResponse(forbit_html)

        if (
            model == "PaymentRecord"
            or model == "SupplyPaymentRecord"
            or model == "SubJobPaymentRecord"
        ):
            # check if the use have the right to modify approval status
            forbit_html = decide_permission(
                request, "approve", {"model": model, "project_id": project_id}
            )
            lock = request.POST.get("lock")
            if lock and forbit_html:
                message = "Bạn không được cấp quyền khóa phiếu. \n\n Vui lí liên hệ admin cấp quyền."
                html_message = render_message(
                    request, message=message, message_type="red"
                )
                return HttpResponse(html_message)

        elif (
            model == "VehicleMaintenance"
            or model == "SupplyOrder"
            or model == "SubJobOrder"
        ):
            current_approval_status = (
                model_class.objects.filter(pk=pk).first().approval_status
            )
            form_approval_status = request.POST.get("approval_status")

            # CHECK PERMISSIONS
            forbit_html = decide_permission(request, "approve", {"model": model})
            # Need to update => all users who has update permission can update
            if current_approval_status == "scratch":
                if form_approval_status not in ("scratch", "wait_for_approval"):
                    message = "Chỉ có thể cập nhật dữ liệu và gửi phê duyệt"
                    html_message = render_message(
                        request, message=message, message_type="red"
                    )
                    return HttpResponse(html_message)

            elif current_approval_status == "wait_for_approval":
                if forbit_html:
                    message = "Không thể cập nhật dữ liệu khi đang chờ phê duyệt"
                    html_message = render_message(
                        request, message=message, message_type="red"
                    )
                    return HttpResponse(html_message)
                else:
                    if form_approval_status not in (
                        "wait_for_approval",
                        "need_update",
                        "approved",
                        "rejected",
                    ):
                        message = 'Không thể chọn lại trạng thái "Bảng nháp"'
                        html_message = render_message(
                            request, message=message, message_type="red"
                        )
                        return HttpResponse(html_message)

            elif current_approval_status == "need_update":
                if form_approval_status not in ("need_update", "wait_for_approval"):
                    message = "Chỉ có thể cập nhật dữ liệu và gửi phê duyệt"
                    html_message = render_message(
                        request, message=message, message_type="red"
                    )
                    return HttpResponse(html_message)

            elif current_approval_status == "approved":
                if forbit_html:
                    if form_approval_status in ("approved"):
                        if model == "VehicleMaintenance":
                            # Update status up each VehicleMaintenanceRepairPart
                            vehicle_part_post_ids = request.POST.getlist(
                                "vehicle_part_id"
                            )
                            for vehicle_part_id in vehicle_part_post_ids:
                                vehicle_part = (
                                    VehicleMaintenanceRepairPart.objects.filter(
                                        id=vehicle_part_id
                                    ).first()
                                )
                                # Get fields
                                vehicle_part.received_status = request.POST.get(
                                    f"received_status_{vehicle_part_id}"
                                )
                                vehicle_part.done_status = request.POST.get(
                                    f"done_status_{vehicle_part_id}"
                                )
                                vehicle_part.save()

                        elif model == "SupplyOrder":
                            order = instance
                            # Update status up each SupplyOrderSupply
                            order_supplies = SupplyOrderSupply.objects.filter(
                                supply_order=order
                            )
                            for order_supply in order_supplies:
                                # Update paid and received quantities
                                paid_quantity = float(
                                    request.POST.get(
                                        f"paid_quantity_{order_supply.base_supply.id}",
                                        0,
                                    )
                                )
                                if paid_quantity:
                                    order_supply.paid_quantity = paid_quantity

                                received_quantity = float(
                                    request.POST.get(
                                        f"received_quantity_{order_supply.base_supply.id}",
                                        0,
                                    )
                                )
                                if received_quantity:
                                    order_supply.received_quantity = received_quantity

                                order_supply.save()
                        elif model == "SubJobOrder":
                            order = instance
                            # Update status up each SubJobOrder
                            order_sub_jobs = SubJobOrder.objects.filter(
                                sub_job_order=order
                            )
                            for order_sub_job in order_sub_jobs:
                                # Update paid and received quantities
                                paid_quantity = float(
                                    request.POST.get(
                                        f"paid_quantity_{order_sub_job.base_sub_job.id}",
                                        0,
                                    )
                                )
                                if paid_quantity:
                                    order_sub_job.paid_quantity = paid_quantity

                                received_quantity = float(
                                    request.POST.get(
                                        f"received_quantity_{order_sub_job.base_sub_job.id}",
                                        0,
                                    )
                                )
                                if received_quantity:
                                    order_sub_job.received_quantity = received_quantity

                                order_sub_job.save()

                        record = instance
                        record.save()
                        record.style = "just-updated"
                        html_message = render_message(
                            request, message="Cập nhật thành công"
                        )
                        html_record = render_display_records(
                            request,
                            model=model,
                            records=[record],
                            update="True",
                            project_id=project_id,
                        )
                        return HttpResponse(html_message + html_record)
                    else:
                        return HttpResponse(forbit_html)

                else:
                    if form_approval_status in ("need_update", "approved", "rejected"):
                        if model == "VehicleMaintenance":
                            # Update status up each VehicleMaintenanceRepairPart
                            vehicle_part_post_ids = request.POST.getlist(
                                "vehicle_part_id"
                            )
                            for vehicle_part_id in vehicle_part_post_ids:
                                vehicle_part = (
                                    VehicleMaintenanceRepairPart.objects.filter(
                                        id=vehicle_part_id
                                    ).first()
                                )
                                # Get fields
                                vehicle_part.received_status = request.POST.get(
                                    f"received_status_{vehicle_part_id}"
                                )
                                vehicle_part.paid_status = request.POST.get(
                                    f"paid_status_{vehicle_part_id}"
                                )
                                vehicle_part.done_status = request.POST.get(
                                    f"done_status_{vehicle_part_id}"
                                )
                                vehicle_part.save()

                        elif model == "SupplyOrder":
                            order = instance

                            # Update status up each SupplyOrderSupply
                            order_supplies = SupplyOrderSupply.objects.filter(
                                supply_order=order
                            )
                            for order_supply in order_supplies:
                                # Update paid and received quantities
                                paid_quantity = float(
                                    request.POST.get(
                                        f"paid_quantity_{order_supply.base_supply.id}",
                                        0,
                                    )
                                )
                                order_supply.paid_quantity = paid_quantity

                                received_quantity = float(
                                    request.POST.get(
                                        f"received_quantity_{order_supply.base_supply.id}",
                                        0,
                                    )
                                )
                                order_supply.received_quantity = received_quantity

                                order_supply.save()

                        elif model == "SubJobOrder":
                            order = instance
                            # Update status up each SubJobOrderSubJob
                            order_sub_jobs = SubJobOrderSubJob.objects.filter(
                                sub_job_order=order
                            )
                            for order_sub_job in order_sub_jobs:
                                # Update paid and received quantities
                                paid_quantity = float(
                                    request.POST.get(
                                        f"paid_quantity_{order_sub_job.base_sub_job.id}",
                                        0,
                                    )
                                )
                                order_sub_job.paid_quantity = paid_quantity

                                received_quantity = float(
                                    request.POST.get(
                                        f"received_quantity_{order_sub_job.base_sub_job.id}",
                                        0,
                                    )
                                )
                                order_sub_job.received_quantity = received_quantity

                                order_sub_job.save()

                            record = instance
                            record.approval_status = form_approval_status
                            record.save()
                            record.style = "just-updated"
                            html_message = render_message(
                                request, message="Cập nhật thành công"
                            )
                            html_record = render_display_records(
                                request,
                                model=model,
                                records=[record],
                                update="True",
                                project_id=project_id,
                            )
                            return HttpResponse(html_message + html_record)

                    else:
                        message = 'Chỉ có thể chọn trạng thái duyệt "Cần sửa lại" hoặc "Từ chối"'
                        html_message = render_message(
                            request, message=message, message_type="red"
                        )
                        return HttpResponse(html_message)
            elif current_approval_status == "rejected":
                if forbit_html:
                    return HttpResponse(forbit_html)
                else:
                    if form_approval_status not in ("need_update"):
                        message = 'Chỉ có thể chọn trạng thái duyệt "Cần sửa lại"'
                        html_message = render_message(
                            request, message=message, message_type="red"
                        )
                        return HttpResponse(html_message)
                    else:
                        record = instance
                        record.approval_status = form_approval_status
                        record.save()
                        record.style = "just-updated"
                        html_message = render_message(
                            request, message="Cập nhật thành công"
                        )
                        html_record = render_display_records(
                            request,
                            model=model,
                            records=[record],
                            update="True",
                            project_id=project_id,
                        )
                        return HttpResponse(html_message + html_record)

    if form.is_valid():
        instance_form = form.save(commit=False)

        if model == "VehicleMaintenance":
            print("\n\n>>>>>>>>>>>>>>", instance_form.id)
            instance_form.save()
            # Update the  list of VehicleMaintenanceRepairPart
            vehicle_maintenance = instance_form
            # get the list of vehicle_parts VehicleMaintenanceRepairPart
            vehicle_parts = VehicleMaintenanceRepairPart.objects.filter(
                vehicle_maintenance=vehicle_maintenance
            )
            part_ids = request.POST.getlist("part_id")

            # check if the id is in the list, if not delete it
            for vehicle_part in vehicle_parts:
                if str(vehicle_part.repair_part.id) not in part_ids:
                    vehicle_part.delete()

            for part_id in part_ids:
                # get the instance VehicleMaintenanceRepairPart which has the repair_part.part_id == part_id
                part = RepairPart.objects.filter(id=part_id).first()
                if part:
                    vehicle_part = VehicleMaintenanceRepairPart.objects.filter(
                        vehicle_maintenance=vehicle_maintenance, repair_part=part
                    ).first()
                # Update
                if vehicle_part:  # Update quantity
                    vehicle_part.quantity = request.POST.get(f"part_quantity_{part_id}")
                    vehicle_part.save()
                else:  # create new
                    if part:
                        VehicleMaintenanceRepairPart.objects.create(
                            vehicle_maintenance=vehicle_maintenance,
                            repair_part=part,
                            quantity=request.POST.get(f"part_quantity_{part_id}"),
                        )

        elif model == "SupplyOrder":
            instance_form.save()
            order = instance_form
            # get the list of vehicle_parts VehicleMaintenanceRepairPart
            order_supplies = SupplyOrderSupply.objects.filter(supply_order=order)
            supply_ids = request.POST.getlist("supply_id")

            # check if the id is in the list, if not delete it
            for order_supply in order_supplies:
                if str(order_supply.base_supply.id) not in supply_ids:
                    order_supply.delete()

            for supply_id in supply_ids:
                # get the instance VehicleMaintenanceRepairPart which has the repair_part.part_id == part_id
                supply = BaseSupply.objects.filter(id=supply_id).first()
                if supply:
                    order_supply = SupplyOrderSupply.objects.filter(
                        supply_order=order, base_supply=supply
                    ).first()
                # Update
                if order_supply:  # Update quantity
                    quantity = float(
                        request.POST.get(f"supply_quantity_{supply_id}", 0)
                    )
                    order_supply.quantity = quantity

                    # add detail supply
                    detail_supply_id = get_valid_id(
                        request.POST.get(f"detail_supply_{supply_id}")
                    )
                    detail_supply = DetailSupply.objects.filter(
                        id=detail_supply_id
                    ).first()
                    if detail_supply:
                        order_supply.detail_supply = detail_supply

                    order_supply.save()

                else:  # create new
                    if supply:
                        quantity = float(
                            request.POST.get(f"supply_quantity_{supply_id}", 0)
                        )
                        SupplyOrderSupply.objects.create(
                            supply_order=order,
                            base_supply=supply,
                            quantity=quantity,
                        )

            # Update approval status if the approval status has changed
            if order.approval_status == "wait_for_approval":

                # Check if order_supplies  count > 0
                order_supplies = SupplyOrderSupply.objects.filter(supply_order=order)
                if order_supplies.count() == 0:
                    record = instance_form
                    record.style = "just-updated"
                    record.approval_status = "scratch"
                    record.save()
                    html_record = render_display_records(
                        request,
                        model=model,
                        records=[record],
                        update="True",
                        project_id=project_id,
                    )
                    html_message = render_message(
                        request,
                        message="Cập nhật thành công nhưng không thể gửi duyệt đơn mua vật tư này vì không có vật tư nào được chọn",
                    )
                    return HttpResponse(html_message + html_record)
                else:
                    for order_supply in order_supplies:
                        if not order_supply.detail_supply:
                            record = instance_form
                            record.style = "just-updated"
                            record.approval_status = "scratch"
                            record.save()
                            html_record = render_display_records(
                                request,
                                model=model,
                                records=[record],
                                update="True",
                                project_id=project_id,
                            )
                            html_message = render_message(
                                request,
                                message="Cập nhật thành công nhưng không thể gửi duyệt đơn mua này vì có vật tư chưa được chọn nhà cung cấp.\n\n Vui lòng liên hệ kế toán để thêm nhà cung cấp.",
                            )
                            return HttpResponse(html_message + html_record)

        elif model == "SubJobOrder":
            instance_form.save()
            order = instance_form
            # get the list of sub_job_order_sub_job records for this project and sub_job
            order_sub_jobs = SubJobOrderSubJob.objects.filter(sub_job_order=order)
            sub_job_ids = request.POST.getlist("sub_job_id")

            # check if the id is in the list, if not delete it
            for order_sub_job in order_sub_jobs:
                if str(order_sub_job.base_sub_job.id) not in sub_job_ids:
                    order_sub_job.delete()

            for sub_job_id in sub_job_ids:
                # get the instance SubJobOrderSubJob which has the sub_job.id == sub_job_id
                sub_job = BaseSubJob.objects.filter(id=sub_job_id).first()
                if sub_job:
                    order_sub_job = SubJobOrderSubJob.objects.filter(
                        sub_job_order=order, base_sub_job=sub_job
                    ).first()
                # Update
                if order_sub_job:  # Update quantity
                    quantity = float(
                        request.POST.get(f"sub_job_quantity_{sub_job_id}", 0)
                    )
                    order_sub_job.quantity = quantity

                    # add detail sub_job
                    detail_sub_job_id = get_valid_id(
                        request.POST.get(f"detail_sub_job_{sub_job_id}")
                    )

                    detail_sub_job = DetailSubJob.objects.filter(
                        id=detail_sub_job_id
                    ).first()
                    if detail_sub_job:
                        order_sub_job.detail_sub_job = detail_sub_job

                    # add sub_job_price
                    sub_job_price = int(
                        request.POST.get(f"sub_job_price_{sub_job_id}", 0)
                    )
                    order_sub_job.sub_job_price = sub_job_price

                    order_sub_job.save()

                else:  # create new
                    if sub_job:
                        quantity = float(
                            request.POST.get(f"sub_job_quantity_{sub_job_id}", 0)
                        )
                        SubJobOrderSubJob.objects.create(
                            sub_job_order=order,
                            base_sub_job=sub_job,
                            quantity=quantity,
                        )
            # Update approval status if the approval status has changed
            if order.approval_status == "wait_for_approval":
                # Check if order_sub_jobs  count > 0
                order_sub_jobs = SubJobOrderSubJob.objects.filter(sub_job_order=order)
                if order_sub_jobs.count() == 0:
                    record = instance_form
                    record.style = "just-updated"
                    record.approval_status = "scratch"
                    record.save()
                    html_record = render_display_records(
                        request,
                        model=model,
                        records=[record],
                        update="True",
                        project_id=project_id,
                    )
                    html_message = render_message(
                        request,
                        message="Cập nhật thành công nhưng không thể gửi duyệt đơn mua vật tư này vì không có vật tư nào được chọn",
                    )
                    return HttpResponse(html_message + html_record)
                else:
                    for order_sub_job in order_sub_jobs:
                        if not order_sub_job.detail_sub_job:
                            record = instance_form
                            record.style = "just-updated"
                            record.approval_status = "scratch"
                            record.save()
                            html_record = render_display_records(
                                request,
                                model=model,
                                records=[record],
                                update="True",
                                project_id=project_id,
                            )
                            html_message = render_message(
                                request,
                                message="Cập nhật thành công nhưng không thể gửi duyệt đơn mua này vì có vật tư chưa chọn nhà cung cấp.\n\n Vui lòng liên hệ kế toán để thêm nhà cung cấp.",
                            )
                            return HttpResponse(html_message + html_record)

        instance_form.save()

        # Save the many to many field, if any
        # form.save_m2m()
        record = instance_form
        record.style = "just-updated"
        html_message = render_message(request, message="Cập nhật thành công")
        html_record = render_display_records(
            request, model=model, records=[record], update="True", project_id=project_id
        )
        return HttpResponse(html_message + html_record)
    else:
        print(form.errors)
        html_modal = render_form(
            request, model=model, pk=pk, form=form, project_id=project_id
        )
        return HttpResponse(html_modal)


@login_required
def get_gantt_chart_data(request, project_id):
    # Get project
    project = get_object_or_404(Project, pk=project_id)
    jobs = project.job_set.all()

    # Sort
    jobs = filter_records(request, jobs, Job)

    # Get checkdate from params
    check_date = request.GET.get("check_date")

    # Return json data including job names, start and end dates
    data = []
    for job in jobs:
        data.append(
            {
                "id": job.secondary_id,
                "name": job.name,
                "start": job.start_date.isoformat(),
                "end": job.end_date.isoformat(),
                "progress_time": progress_by_time(job, check_date=check_date)[
                    "percent"
                ],
                "progress_amount": progress_by_amount(job, check_date=check_date)[
                    "percent"
                ],
            }
        )

    return JsonResponse(data, safe=False)


@login_required
def load_elements(request):
    encoded_params = request.GET.get("q", "")
    params = json.loads(decode_params(encoded_params))
    for key, value in request.GET.items():
        if key != "q":
            params[key] = value
    print("\n>>>>>>>>>> elements params:", params)
    html = '<div id="load-elements" class"hidden"></div>'

    # CHECK PERMISSIONS
    forbit_html = decide_permission(request, "read", params)
    if forbit_html:
        return HttpResponse(forbit_html)

    # RENDER ELEMENTS
    elements = params.get("elements", "")
    for element in elements.split("|"):
        element = element.strip()
        if element == "title_bar":
            html_title_bar = render_title_bar(request, **params)
            html += html_title_bar
        elif element == "tool_bar":
            html_tool_bar = render_tool_bar(request, **params)
            html += html_tool_bar
        elif element == "infor_bar":
            html_infor_bar = render_infor_bar(request, **params)
            html += html_infor_bar
        elif element == "display_records":
            html_display_records = render_display_records(request, **params)
            html += html_display_records
        elif element == "message":
            html_message = render_message(request, **params)
            html += html_message
        elif element == "modal_form":
            html_modal_form = render_form(request, **params)
            html += html_modal_form
        elif element == "gantt_chart":
            pass
        elif element == "weekplan_table":
            project_id = get_valid_id(params.get("project_id", 0))
            check_date = get_valid_date(params.get("check_date", ""))
            html_weekplan_table = render_weekplan_table(request, project_id, check_date)
            html += html_weekplan_table

    return HttpResponse(html)


@login_required
def load_weekplan_table(request, project_id):
    check_date = request.GET.get("check_date")
    html_weekplan_table = render_weekplan_table(request, project_id, check_date)
    html_tool_bar = '<div id="tool-bar" class="hidden"></div>'
    return HttpResponse(html_weekplan_table + html_tool_bar)


@login_required
def handle_weekplan_form(request):
    if request.method != "POST":
        return HttpResponseForbidden()
    form = request.POST

    try:
        start_date = form.get("start_date")
        end_date = form.get("end_date")
        check_date = form.get("check_date")
        project_id = form.get("project_id")
        project = Project.objects.get(pk=project_id)

        # get all jobs of the project
        jobs = Job.objects.filter(project_id=project_id)

        weekplan_status = form.get("weekplan_status")
        user_role = ProjectUser.objects.filter(
            project=project, user=request.user
        ).first()

        if user_role.role == "technician" or user_role.role == "all":
            message = "Phê duyệt thành công"
            for job in jobs:
                jobplan = JobPlan.objects.filter(
                    job=job, start_date=start_date, end_date=end_date
                ).first()
                if jobplan:
                    jobplan.status = weekplan_status
                    jobplan.save()

        elif user_role.role == "supervisor" or user_role.role == "all":
            message = "Cập nhật kế hoạch tuần thành công. \n\nĐã chuyển kế hoạch đến người xét duyệt!"
            weekplan_status = "wait_for_approval"
            for job in jobs:
                note = form.get(f"note_{job.pk}")
                quantity = form.get(f"plan_quantity_{job.pk}")
                try:
                    quantity = float(quantity)
                except ValueError as e:
                    quantity = 0

                if type(note) != str:
                    note = ""

                if quantity > job.quantity:
                    message = f'Lỗi khi nhập khối lượng kế hoạch cho công việc "{job.name}". Khối lượng kế hoạch phải nhỏ hơn khối lượng công việc ({str(job.quantity)}).'
                    html_message = render_message(
                        request, message=message, message_type="red"
                    )
                    # print(html_message  )
                    return HttpResponse(html_message)

                jobplan = JobPlan.objects.filter(
                    job=job, start_date=start_date, end_date=end_date
                ).first()
                if quantity == 0 and note.strip() == "":
                    if jobplan:
                        jobplan.delete()
                        pass
                    continue

                if jobplan:
                    jobplan.note = note
                    jobplan.plan_quantity = quantity
                    jobplan.status = weekplan_status
                    jobplan.save()
                    continue
                else:
                    JobPlan(
                        job=job,
                        start_date=start_date,
                        end_date=end_date,
                        plan_quantity=quantity,
                        note=note,
                        status=weekplan_status,
                    ).save()

        html_weekplan_table = render_weekplan_table(
            request, project_id, check_date=check_date
        )
        html_infor_bar = render_infor_bar(
            request,
            page="page_each_project",
            project_id=project_id,
            check_date=check_date,
        )
        html_message = render_message(request, message=message)
        return HttpResponse(html_message + html_weekplan_table + html_infor_bar)

    except Exception as e:
        raise e
        html = render_message(request, message="Có lỗi: " + str(e), message_type="red")
        return HttpResponse(html)


@login_required
def handle_date_report_form(request):
    if request.method != "POST":
        return HttpResponseForbidden()

    form = request.POST
    check_date = form.get("check_date")
    try:
        check_date = datetime.strptime(check_date, "%Y-%m-%d").date()
    except:
        check_date = timezone.now().date()
    # Get monday and sunday dates of the week that contains check_date
    monday = check_date - timedelta(days=check_date.weekday())
    sunday = check_date + timedelta(days=6 - check_date.weekday())

    try:
        check_date = form.get("check_date")
        project_id = form.get("project_id")
        # get all jobs of the project
        jobs = Job.objects.filter(project_id=project_id)

        has_jobplan = False
        for job in jobs:
            jobplan_in_week = JobPlan.objects.filter(
                start_date__gte=monday, end_date__lte=sunday, job=job
            ).first()
            if jobplan_in_week:
                if jobplan_in_week.status == "wait_for_approval":
                    html = render_message(
                        request,
                        message="Không cập nhật được báo cáo ngày.\n\nVui lòng chờ báo cáo tuần được phê duyệt trước.",
                        message_type="red",
                    )
                    return HttpResponse(html)
                elif jobplan_in_week.status == "rejected":
                    html = render_message(
                        request,
                        message="Không cập nhật được báo cáo ngày.\n\nBáo cáo tuần đã bị từ chối, vui lòng cập nhật lại báo cáo tuần.",
                        message_type="red",
                    )
                    return HttpResponse(html)
                elif jobplan_in_week.status == "approved":
                    has_jobplan = True
        if not has_jobplan:
            html = render_message(
                request,
                message="Không cập nhật được báo cáo ngày.\n\nChưa tạo báo cáo tuần.",
                message_type="red",
            )
            return HttpResponse(html)

        for job in jobs:
            note = form.get(f"date_note_{job.pk}")
            quantity = form.get(f"date_quantity_{job.pk}")
            material = form.get(f"date_material_{job.pk}")
            labor = form.get(f"date_labor_{job.pk}")
            try:
                quantity = float(quantity)
            except:
                quantity = 0

            if type(note) != str:
                note = ""
            if type(material) != str:
                material = ""
            if type(labor) != str:
                labor = ""

            job_date_report = JobDateReport.objects.filter(
                job=job, date=check_date
            ).first()

            current_date_quantity = job_date_report.quantity if job_date_report else 0

            total_quantity_reported = (
                progress_by_quantity(job)["total_quantity_reported"]
                - current_date_quantity
            )
            total_quantity_left = job.quantity - total_quantity_reported

            if int(quantity) > total_quantity_left:
                message = f'Lỗi khi nhập khối lượng hoàn thành (đang nhập {quantity}) trong ngày cho công việc "{job.name}". Khối lượng hoàn thành phải nhỏ hơn khối lượng còn lại ({str(total_quantity_left)}).'
                html_message = render_message(
                    request, message=message, message_type="red"
                )
                return HttpResponse(html_message)

            if (
                quantity == 0
                and note.strip() == ""
                and material.strip() == ""
                and labor.strip() == ""
            ):
                if job_date_report:
                    job_date_report.delete()
                continue

            if job_date_report:
                job_date_report.note = note
                job_date_report.quantity = quantity
                job_date_report.material = material
                job_date_report.labor = labor
                job_date_report.save()
                continue
            else:
                JobDateReport(
                    job=job,
                    date=check_date,
                    quantity=quantity,
                    material=material,
                    labor=labor,
                    note=note,
                ).save()

        html_weekplan_table = render_weekplan_table(
            request, project_id, check_date=check_date
        )
        html_message = render_message(request, message="Cập nhật báo cáo thành công")
        html_infor_bar = render_infor_bar(
            request,
            page="page_each_project",
            project_id=project_id,
            check_date=check_date,
        )
        return HttpResponse(html_message + html_weekplan_table + html_infor_bar)

    except Exception as e:
        raise e
        html = render_message(request, message="Có lỗi: " + str(e), message_type="red")
        return HttpResponse(html)


def handle_vehicle_operation_form(request):
    def convert_time(time_str, time_sign):
        hours, minutes, seconds = map(int, time_str.split(":"))
        total_seconds = hours * 3600 + minutes * 60 + seconds
        duration_seconds = total_seconds
        if time_sign == "minus":
            duration_seconds = -duration_seconds
        return duration_seconds

    if request.method != "POST":
        return HttpResponseForbidden()

    try:
        form = request.POST
        # Get list of ids
        ids = form.getlist("id")
        for id in ids:
            driver = StaffData.objects.filter(
                pk=get_valid_id(form.get(f"driver_{id}", None))
            ).first()
            location = Location.objects.filter(
                pk=get_valid_id(form.get(f"location_{id}", None))
            ).first()
            record = VehicleOperationRecord.objects.get(pk=id)
            record.driver = driver
            record.location = location
            record.fuel_allowance = form.get(f"fuel_allowance_{id}", None)
            record.note = form.get(f"note_{id}", None)
            record.allow_overtime = (
                False if not form.get(f"allow_overtime_{id}", False) else True
            )
            if record.source == "manual":
                try:
                    # duration seconds
                    duration_seconds_str = form.get(f"duration_seconds_{id}", None)
                    duration_seconds_sign = form.get(
                        f"duration_seconds_sign_{id}", None
                    )
                    duration_seconds = convert_time(
                        duration_seconds_str, duration_seconds_sign
                    )
                    record.duration_seconds = duration_seconds
                    # over time
                    overtime_str = form.get(f"overtime_{id}", None)
                    overtime_sign = form.get(f"overtime_sign_{id}", None)
                    overtime = convert_time(overtime_str, overtime_sign)
                    record.overtime = overtime
                    # normal_woring_time
                    normal_working_time_str = form.get(
                        f"normal_working_time_{id}", None
                    )
                    normal_working_time_sign = form.get(
                        f"normal_working_time_sign_{id}", None
                    )
                    normal_working_time = convert_time(
                        normal_working_time_str, normal_working_time_sign
                    )
                    record.normal_working_time = normal_working_time
                except:
                    raise e

                if not driver:
                    # remove record
                    record.delete()
                    continue
            record.save()

        # Add new records
        new_index = 0
        while True:
            new_index += 1
            if not form.get(f"vehicle_new_{new_index}", None):
                break

            vehicle_new = form.get(f"vehicle_new_{new_index}", None)
            start_time_new = form.get(f"start_time_new_{new_index}", None)
            driver_new = form.get(f"driver_new_{new_index}", None)
            location_new = form.get(f"location_new_{new_index}", None)
            duration_seconds_new = form.get(f"duration_seconds_new_{new_index}", None)
            duration_seconds_sign_new = form.get(
                f"duration_seconds_sign_new_{new_index}", None
            )
            overtime_new = form.get(f"overtime_new_{new_index}", None)
            overtime_sign_new = form.get(f"overtime_sign_new_{new_index}", None)
            normal_working_time_new = form.get(
                f"normal_working_time_new_{new_index}", None
            )
            normal_working_time_sign_new = form.get(
                f"normal_working_time_sign_new_{new_index}", None
            )
            fuel_allowance_new = form.get(f"fuel_allowance_new_{new_index}", None)
            note_new = form.get(f"note_new_{new_index}", None)
            allow_overtime_new = (
                False
                if not form.get(f"allow_overtime_new_{new_index}", False)
                else True
            )
            try:
                # duration seconds
                duration_seconds_str = duration_seconds_new
                duration_seconds_sign = duration_seconds_sign_new
                duration_seconds = convert_time(
                    duration_seconds_str, duration_seconds_sign
                )

                # over time
                overtime_str = overtime_new
                overtime_sign = overtime_sign_new
                overtime = convert_time(overtime_str, overtime_sign)

                # normal_woring_time
                normal_working_time_str = normal_working_time_new
                normal_working_time_sign = normal_working_time_sign_new
                normal_working_time = convert_time(
                    normal_working_time_str, normal_working_time_sign
                )

            except Exception as e:
                raise e

            vehicle = vehicle_new
            driver = StaffData.objects.filter(pk=get_valid_id(driver_new)).first()
            location = Location.objects.filter(pk=get_valid_id(location_new)).first()

            start_time = start_time_new
            # convert start time to datetime object
            start_time = datetime.strptime(start_time, "%d/%m/%Y")
            end_time = start_time
            fuel_allowance = get_valid_int(fuel_allowance_new)
            note = note_new
            allow_overtime = allow_overtime_new

            if not driver:
                continue

            new_record = VehicleOperationRecord.objects.create(
                vehicle=vehicle,
                start_time=start_time,
                end_time=end_time,
                duration_seconds=duration_seconds,
                overtime=overtime,
                normal_working_time=normal_working_time,
                fuel_allowance=fuel_allowance,
                note=note,
                source="manual",
                driver=driver,
                location=location,
                allow_overtime=allow_overtime,
            )
            ids.append(new_record.id)

        # Get records by ids
        group_by = form.get("group_by")
        tab = form.get("tab")
        records = VehicleOperationRecord.objects.filter(pk__in=ids).order_by(
            "source", "start_time"
        )

        # Get start_date from start_time

        start_date = records.first().start_time.date()
        end_date = records.last().end_time.date()
        html_display = render_display_records(
            request,
            model="VehicleOperationRecord",
            start_date=start_date,
            end_date=end_date,
            records=records,
            group_by=group_by,
            tab=tab,
            update=True,
        )

        html_message = render_message(
            request,
            message="Cập nhật thành công!\n\nLưu ý các dòng nhập tay nếu không có TÀI XẾ sẽ bị xóa",
            message_type="green",
        )
        html = html_message + html_display
        return HttpResponse(html)
    except Exception as e:
        html = render_message(request, message="Có lỗi: " + str(e), message_type="red")
        return HttpResponse(html)


@login_required
def download_excel_template(request, template_name):
    if template_name == "jobs":
        # Get the Excel file from media/excel/Mẫu công việc trong dự án.xlsx
        excel_file = os.path.join(
            settings.STATIC_ROOT, "excel", f"cong-viec-trong-du-an.xlsx"
        )
        # Return the file
        with open(excel_file, "rb") as f:
            excel_data = f.read()
        response = HttpResponse(
            excel_data,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="cong-viec-trong-du-an.xlsx"'
        )

    elif template_name == "cost_estimation_table":
        original_excel_path = os.path.join(
            settings.STATIC_ROOT, "excel", f"mau-bang-du-toan-vat-tu.xlsx"
        )

        # Create a temporary copy of the file to avoid race conditions
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            temp_excel_path = temp_file.name
            shutil.copy(original_excel_path, temp_excel_path)  # Copy original file

        # Load the copied file (not the original) to preserve formatting
        wb = load_workbook(temp_excel_path)

        # Modify or create the "data" sheet
        if "Vật tư" in wb.sheetnames:
            sheet = wb["Vật tư"]
            sheet.delete_rows(2, sheet.max_row)  # Clear existing data (except headers)
        else:
            sheet = wb.create_sheet("Vật tư")

        # Get all data from BasicSupply
        records = BaseSupply.objects.all()
        # convert to data
        supplies = []
        for record in records:
            supplies.append(
                [
                    record.supply_number + " - " + record.supply_name,
                    record.unit,
                    "#" + record.supply_number,
                    record.material_type,
                ]
            )

        # Append new data to the "data" sheet
        for row in supplies:
            sheet.append(row)

        # Save modifications to the temporary file
        wb.save(temp_excel_path)

        # Open the modified file and return it as a response
        with open(temp_excel_path, "rb") as f:
            response = HttpResponse(
                f.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = (
                'attachment; filename="mau-bang-du-toan-vat-tu.xlsx"'
            )

        # Cleanup: Delete the temporary file after sending the response
        os.remove(temp_excel_path)

    elif template_name == "sub_job_cost_estimation_table":
        original_excel_path = os.path.join(
            settings.STATIC_ROOT, "excel", f"mau-bang-du-toan-cong-viec-thau-phu.xlsx"
        )
        # Create a temporary copy of the file to avoid race conditions
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            temp_excel_path = temp_file.name
            shutil.copy(original_excel_path, temp_excel_path)  # Copy original file

        # Load the copied file (not the original) to preserve formatting
        wb = load_workbook(temp_excel_path)

        # Modify or create the "data" sheet
        if "Công việc" in wb.sheetnames:
            sheet = wb["Công việc"]
            sheet.delete_rows(2, sheet.max_row)  # Clear existing data (except headers)
        else:
            sheet = wb.create_sheet("Công việc")

        # Get all data from BaseSubJob
        records = BaseSubJob.objects.all()
        # convert to data
        sub_jobs = []
        for record in records:
            sub_jobs.append(
                [
                    record.job_number + " " + record.job_name,
                    record.unit,
                    "#" + record.job_number,
                    record.job_type,
                ]
            )

        # Write data to the sheet
        for row in sub_jobs:
            sheet.append(row)

        # Save modifications to the temporary file
        wb.save(temp_excel_path)

        # Open the modified file and return it as a response
        with open(temp_excel_path, "rb") as f:
            response = HttpResponse(
                f.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = (
                'attachment; filename="mau-bang-du-toan-cong-viec-thau-phu.xlsx"'
            )

        os.remove(temp_excel_path)

    return response


@login_required
def upload_project(request, project_id):

    if request.method != "POST":
        return "API này chỉ dùng POST"

    excel_file = request.FILES.get("file")
    project = Project.objects.filter(pk=project_id).first()

    if not excel_file:
        html_message = render_message(
            request, message="Vui lý nhập file Excel", message_type="red"
        )
        return HttpResponse(html_message)
    if not project:
        html_message = render_message(
            request, message="Dự án này không tồn tại", message_type="red"
        )
        return HttpResponse(html_message)

    # Read the data from the Excel file then save as job record
    df = pd.read_excel(excel_file, header=1)

    # Check if header is in the second row
    required_headers = [
        "STT",
        "Danh mục",
        "Tên công việc",
        "Mô tả",
        "Trạng thái",
        "Bắt đầu",
        "Kết thúc",
        "Đơn vị",
        "Khối lượng",
        "Đơn giá",
        "Mã hiệu đơn giá",
    ]

    missing_headers = [
        header for header in required_headers if header not in df.columns
    ]
    if missing_headers:
        html_message = render_message(
            request,
            message=f"File Excel không đúng mẫu. Thiếu các cột: {', '.join(missing_headers)}.\nLưu ý: Tên cột phải ở dòng số 2.",
            message_type="red",
        )
        return HttpResponse(html_message)

    # The table uses verbose names in the excel file, so we need to convert the verbose names to real names
    # Loop through the fields in job
    for field in Job._meta.get_fields():
        if not hasattr(field, "verbose_name"):
            continue
        if field.verbose_name in df.columns:
            df.rename(columns={field.verbose_name: field.name}, inplace=True)
    # remove duplicate with same name and category
    df = df.drop_duplicates(subset=["name", "category"], keep="first")

    # Before saving any data, we must check if the data of each column is valid
    errors = ""
    jobs = []
    for index, row in df.iterrows():
        job = Job()
        job.project = project
        for field in df.columns:
            value = row[field]
            # if value is NaN
            if pd.isna(value):
                value = ""

            # If field has choice field, we need to convert the string to the correct value
            if field == "status":
                real_status = ""
                for choice in job.STATUS_CHOICES:
                    if value == choice[1]:
                        real_status = choice[0]
                        break
                if real_status:
                    setattr(job, field, real_status)
                else:
                    setattr(job, field, value)  # to raise error
            else:
                setattr(job, field, value)
        try:
            job.clean()
        except ValidationError as e:
            errors += f"Hàng {str(index + 1)}:\n {e.message}" + "\n"
        jobs.append(job)
    if errors:
        html_message = render_message(request, message=errors, message_type="red")
        return HttpResponse(html_message)
    else:
        for job in jobs:
            job.save()
        html_message = render_message(
            request, message="Cập nhật thành công", ok_button_function="reload"
        )
        return HttpResponse(html_message)


def get_binhanh_service_operation_time(check_date, vehicles):
    # if check_date is None:
    #     check_date = datetime.now()

    # # convert checkdate to string dd//mm/yyyy
    # check_date_str = check_date.strftime("%d/%m/%Y")

    # Get json data
    def call_api(url, payload, auth):
        response = requests.post(url, json=payload, auth=auth)
        return response

    def get_vehicle_list():
        # get api type from params

        customer_code = "71735_6"
        api_key = "Ff$BkG1rAu"
        auth = HTTPBasicAuth(customer_code, api_key)
        url = "http://api.gps.binhanh.vn/apiwba/gps/tracking"
        payload = {"IsFuel": True}
        response = call_api(url, payload, auth)
        if response.status_code == 200:
            data = response.json()
            message_result = data.get("MessageResult")
            if message_result == "Success":
                vehicles = data.get("Vehicles", [])
                # Extracting PrivateCode values
                private_codes = [vehicle["PrivateCode"] for vehicle in vehicles]
                return private_codes
            else:
                return []
        else:
            return []

    def get_operation_time(vehicles, start_date, end_date):
        # URL for login
        url = "https://gps.binhanh.vn"

        # Start a session to persist cookies across requests
        session = requests.Session()

        # Headers for the request
        headers = {
            "Content-Type": "application/x-www-form-urlencoded",
            "DNT": "1",
            "Origin": "https://gps.binhanh.vn",
            "Referer": "https://gps.binhanh.vn/",
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
            "sec-ch-ua": '"Chromium";v="130", "Google Chrome";v="130", "Not?A_Brand";v="99"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"macOS"',
        }

        # Step 1: Get the initial login page to retrieve `__VIEWSTATE`, `__VIEWSTATEGENERATOR`, and `__EVENTVALIDATION`
        response = session.get(url, headers=headers)
        soup = BeautifulSoup(response.text, "html.parser")

        # Extract the dynamic fields
        viewstate = soup.find("input", {"name": "__VIEWSTATE"})["value"]
        viewstate_generator = soup.find("input", {"name": "__VIEWSTATEGENERATOR"})[
            "value"
        ]
        event_validation = soup.find("input", {"name": "__EVENTVALIDATION"})["value"]

        # Step 2: Prepare the payload for login
        data = {
            "__LASTFOCUS": "",
            "__EVENTTARGET": "",
            "__EVENTARGUMENT": "",
            "__VIEWSTATE": viewstate,
            "__VIEWSTATEGENERATOR": viewstate_generator,
            "__EVENTVALIDATION": event_validation,
            "UserLogin1$txtLoginUserName": "tinnghiavt",
            "UserLogin1$txtLoginPassword": "Tinnghia1234",
            "UserLogin1$hdfPassword": "",
            "UserLogin1$btnLogin": "Đăng nhập",
            "UserLogin1$txtPhoneNumberOtp": "",
            "UserLogin1$txtOTPClient": "",
            "UserLogin1$hdfOTPServer": "",
            "UserLogin1$hdfTimeoutOTP": "",
        }

        # Step 3: Send the POST request to login
        login_response = session.post(url, headers=headers, data=data)
        # Step 4: Check if login was successful by verifying redirection or specific content in the response 01/05/2024
        if login_response.ok and "OnlineM.aspx" in login_response.url:
            operation_time = {}
            count = 0
            for vehicle in vehicles:
                url = "https://gps.binhanh.vn/HttpHandlers/RouteHandler.ashx?method=getRouterByCarNumberLite&carNumber={}&fromDate={}%2000:00&toDate={}%2023:59&split=false&isItinerary=false".format(
                    vehicle, start_date, end_date
                )
                response = session.get(url, headers=headers)
                data = response.json().get("data")
                count += 1
                # print('Vehicle {}/{}:'.format(count, len(vehicles)), vehicle)

                if data == []:
                    operation_time[vehicle] = {}
                    continue
                df = pd.DataFrame(data)
                # Select only columns 1 and 17 (index-based selection)
                df = df.iloc[:, [1, 18]]
                # Rename columns
                df.columns = ["timestamp", "color"]

                # Convert timestamp column to datetime for easier processing
                df["timestamp"] = pd.to_datetime(
                    df["timestamp"], format="%d/%m/%Y %H:%M:%S"
                )

                # Find the start and end of each consecutive color block
                df["change"] = (df["color"] != df["color"].shift()).cumsum()

                # Group by 'color' and 'change' to get periods of consecutive colors
                summary = (
                    df.groupby(["color", "change"])
                    .agg(
                        start_time=("timestamp", "first"),
                        end_time=("timestamp", "last"),
                    )
                    .reset_index()
                )

                # Filter to only include rows where color is "Blue"
                blue_summary = summary[
                    summary["color"] == "Blue"
                ].copy()  # Make an explicit copy
                # Convert start_time and end_time columns to datetime
                blue_summary["start_time"] = pd.to_datetime(blue_summary["start_time"])
                blue_summary["end_time"] = pd.to_datetime(blue_summary["end_time"])
                # Calculate duration in seconds
                blue_summary["duration_seconds"] = (
                    blue_summary["end_time"] - blue_summary["start_time"]
                ).dt.total_seconds()

                # Convert start_time and end_time to string format for JSON serialization
                blue_summary["start_time"] = blue_summary["start_time"].astype(str)
                blue_summary["end_time"] = blue_summary["end_time"].astype(str)

                operation_time[vehicle] = blue_summary.to_dict(orient="records")
            return operation_time
        else:
            return []

    if vehicles == []:
        vehicles = get_vehicle_list()

    # start_date = '01/05/2024'
    # end_date = '01/05/2024'

    # Get operation time
    operation_time = get_operation_time(vehicles, check_date, check_date)
    return operation_time


def get_vehicle_list_from_binhanh(request):
    def call_api(url, payload, auth):
        response = requests.post(url, json=payload, auth=auth)
        return response

    # get api type from params
    customer_code = "71735_6"
    api_key = "Ff$BkG1rAu"
    auth = HTTPBasicAuth(customer_code, api_key)
    url = "http://api.gps.binhanh.vn/apiwba/gps/tracking"
    payload = {"IsFuel": True}
    response = call_api(url, payload, auth)
    if response.status_code == 200:
        data = response.json()
        message_result = data.get("MessageResult")
        if message_result == "Success":
            vehicles = data.get("Vehicles", [])
            # Extracting PrivateCode values
            private_codes = [vehicle["PrivateCode"] for vehicle in vehicles]
            return JsonResponse(private_codes, safe=False)
        else:
            return []
    else:
        return []


def get_trip_data_from_binhanh(request):

    def parse_operation_time(vehicle, data):
        if settings.DOMAIN == "localhost":
            with open("local/log_api.json", "a") as f:
                f.write("\n\n=========== vehicle: " + vehicle + "===========\n")
                f.write(
                    "=========== run time:"
                    + datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    + "===========\n"
                )
                json.dump(data, f, indent=4)

        operation_time = {}
        routes = data.get("Routes", [])
        if not routes:
            operation_time[vehicle] = {}
            return operation_time

        # Convert the list of routes into a DataFrame
        df = pd.DataFrame(routes)
        # Select only the required columns
        df = df[["LocalTime", "IsMachineOn"]]
        # Convert LocalTime column to datetime for easier processing
        df["LocalTime"] = pd.to_datetime(df["LocalTime"], format="%Y-%m-%dT%H:%M:%S")

        # Find the start and end of each consecutive IsMachineOn block
        df["change"] = (df["IsMachineOn"] != df["IsMachineOn"].shift()).cumsum()

        # Group by 'IsMachineOn' and 'change' to get periods of consecutive colors
        summary = (
            df.groupby(["IsMachineOn", "change"])
            .agg(start_time=("LocalTime", "first"), end_time=("LocalTime", "last"))
            .reset_index()
        )

        # Filter to only include rows where IsMachineOn is "True"
        blue_summary = summary[
            summary["IsMachineOn"] == True
        ].copy()  # Make an explicit copy
        # Convert start_time and end_time columns to datetime
        blue_summary["start_time"] = pd.to_datetime(blue_summary["start_time"])
        blue_summary["end_time"] = pd.to_datetime(blue_summary["end_time"])

        # Calculate duration in seconds
        blue_summary["duration_seconds"] = (
            blue_summary["end_time"] - blue_summary["start_time"]
        ).dt.total_seconds()

        # Convert start_time and end_time to string format for JSON serialization
        blue_summary["start_time"] = blue_summary["start_time"].astype(str)
        blue_summary["end_time"] = blue_summary["end_time"].astype(str)

        operation_time[vehicle] = blue_summary.to_dict(orient="records")
        return operation_time

    def save_operation_record(operation_time):
        # Parse JSON data from the request body
        result = ""
        for vehicle, other_values_list in operation_time.items():
            log_result_data = ""
            for other_values in other_values_list:
                start_time = other_values.get("start_time")
                start_time = datetime.strptime(start_time, "%Y-%m-%d %H:%M:%S")
                end_time = other_values.get("end_time")
                end_time = datetime.strptime(end_time, "%Y-%m-%d %H:%M:%S")
                duration_seconds = other_values.get("duration_seconds")

                # check if the records which has start_time in the check_date
                vehicle_operation_record = VehicleOperationRecord.objects.filter(
                    vehicle=vehicle, start_time=start_time
                ).first()

                # Use this to make sure the vehicle_operation_record is not None
                if not vehicle_operation_record:
                    vehicle_operation_records = VehicleOperationRecord.objects.filter(
                        vehicle=vehicle, start_time__date=start_time.date()
                    )
                    for each_vehicle_operation_record in vehicle_operation_records:
                        if (
                            each_vehicle_operation_record.start_time == start_time
                            and each_vehicle_operation_record.vehicle == vehicle
                        ):
                            vehicle_operation_record = each_vehicle_operation_record

                if vehicle_operation_record:
                    vehicle_operation_record.end_time = end_time
                    vehicle_operation_record.duration_seconds = duration_seconds
                    vehicle_operation_record.save()
                    # log_result_data += f'- Update record with id {vehicle_operation_record.id}: ' + str(vehicle) + ' - ' + str(start_time) + ' - ' + str(end_time) + ' - ' + str(duration_seconds) + '\n'
                else:
                    # Create and save the VehicleOperationRecord instance
                    VehicleOperationRecord.objects.create(
                        vehicle=vehicle,
                        start_time=start_time,
                        end_time=end_time,
                        duration_seconds=duration_seconds,
                    )
                    log_result_data += (
                        f"- Create record: "
                        + str(vehicle)
                        + " - "
                        + str(start_time)
                        + " - "
                        + str(end_time)
                        + " - "
                        + str(duration_seconds)
                        + "\n"
                    )
            result += (
                vehicle + " => Success - Details below:\n" + log_result_data + "\n"
            )
        return result

    # get check_date from url
    gps_name = request.GET.get("gps_name")
    check_date = request.GET.get("check_date")
    # Define the API endpoint
    url = "http://api.gps.binhanh.vn/api/gps/route"

    # Define the payload (parameters)
    # from_date="2024-11-18T00:00:00", to_date="2024-11-18T17:30:00"
    payload = {
        "CustomerCode": "71735_6",  # Replace with your customer code
        "key": "Ff$BkG1rAu",  # Replace with your API key
        "vehiclePlate": gps_name,  # Replace with the vehicle plate
        "fromDate": check_date
        + "T00:00:00",  # Replace with the desired start date and time
        "toDate": check_date
        + "T23:59:59",  # Replace with the desired end date and time
    }

    # Define the headers (optional, if needed)
    headers = {"Content-Type": "application/json"}
    # print('\n>>>> Get data for vehicle: ', gps_name, ' on date: ', check_date)
    try:
        # Make the POST request
        response = requests.post(url, json=payload, headers=headers)
        # Check the status code
        if response.status_code == 200:
            data = response.json()
            # return JsonResponse(data)
            operation_time = parse_operation_time(gps_name, data)
            # return JsonResponse(operation_time, safe=False)
            result = save_operation_record(operation_time)
            return HttpResponse(result)
        else:
            result = "Request failed with status code: " + str(response.status_code)
            result += "\nResponse: " + str(response.text)
            # print(result)
            return HttpResponse(result)

    except requests.exceptions.RequestException as e:
        # print("An error occurred:", e)
        result = "An error occurred: " + str(e)
        return HttpResponse(result)


@csrf_exempt
def save_vehicle_operation_record(request):
    # only accept POST request
    if request.method != "POST":
        return HttpResponse("Method not allowed")

    # get check_date from url
    check_date = request.POST.get("check_date")
    vehicles = request.POST.getlist("vehicle")
    check_date = get_valid_date(check_date)
    # convert check_date to datetime date
    check_date = datetime.strptime(check_date, "%Y-%m-%d").date()
    # convert check_date to ddd/mm/yyyy
    check_date = check_date.strftime("%d/%m/%Y")

    data = get_binhanh_service_operation_time(check_date, vehicles)
    # Parse JSON data from the request body
    result = ""
    for vehicle, other_values_list in data.items():
        for other_values in other_values_list:
            start_time = other_values.get("start_time")
            start_time = datetime.strptime(start_time, "%Y-%m-%d %H:%M:%S")
            end_time = other_values.get("end_time")
            end_time = datetime.strptime(end_time, "%Y-%m-%d %H:%M:%S")
            duration_seconds = other_values.get("duration_seconds")

            # check if the records which has start_time in the check_date
            vehicle_operation_record = VehicleOperationRecord.objects.filter(
                vehicle=vehicle, start_time=start_time
            ).first()

            if vehicle_operation_record:
                vehicle_operation_record.end_time = end_time
                vehicle_operation_record.duration_seconds = duration_seconds
                vehicle_operation_record.save()
            else:
                # Create and save the VehicleOperationRecord instance
                VehicleOperationRecord.objects.create(
                    vehicle=vehicle,
                    start_time=start_time,
                    end_time=end_time,
                    duration_seconds=duration_seconds,
                )
        result += vehicle + "\n"
    return HttpResponse(check_date + " => done")


def form_repair_parts(request):
    providers = PartProvider.objects.all()
    repair_parts = RepairPart.objects.all()
    context = {"repair_parts": repair_parts, "providers": providers}
    return render(request, "components/modal_repair_parts.html", context)


def form_detailed_supplies(request):
    providers = SupplyProvider.objects.all()
    supplies = DetailSupply.objects.all()
    context = {"supplies": supplies, "providers": providers}
    return render(request, "components/modal_detail_supplies.html", context)


def form_base_supplies(request):
    project_id = request.GET.get("project")
    # Get cost estimations for the project
    estimations = CostEstimation.objects.filter(project_id=project_id)
    base_supplies = BaseSupply.objects.filter(
        id__in=estimations.values_list("base_supply_id", flat=True)
    )

    # supplies = BaseSupply.objects.all()
    context = {
        "estimations": estimations,
        "project_id": project_id,
    }
    return render(request, "components/modal_base_supplies.html", context)


def form_base_sub_jobs(request):
    project_id = request.GET.get("project")
    # Get cost estimations for the project
    estimations = SubJobEstimation.objects.filter(project_id=project_id)
    base_sub_jobs = BaseSubJob.objects.filter(
        id__in=estimations.values_list("base_sub_job_id", flat=True)
    )
    context = {
        "estimations": estimations,
        "project_id": project_id,
    }
    return render(request, "components/modal_base_sub_jobs.html", context)


def form_cost_estimation_table(request, project_id):
    def render_modal(project_id, message=None, message_type="green"):

        # Get fields to be displayed by using record meta
        # If there is get_display_fields method, use that method
        fields = []
        headers = []
        if hasattr(CostEstimation, "get_display_fields"):

            for field in CostEstimation.get_display_fields():
                fields.append(field)
                headers.append(CostEstimation._meta.get_field(field).verbose_name)

        # Get the list of repair parts
        records = CostEstimation.objects.filter(project=project_id)
        context = {
            "records": records,
            "fields": fields,
            "headers": headers,
            "project_id": project_id,
            "message": message,
            "message_type": message_type,
        }
        return render(request, "components/modal_cost_estimation_table.html", context)

    if request.method == "GET":
        forbit_html = decide_permission(request, "read", {"model": "CostEstimation"})
        if forbit_html:
            return HttpResponse(forbit_html)
        else:
            return render_modal(project_id)

    if request.method == "POST":
        forbit_html = decide_permission(request, "create", {"model": "CostEstimation"})
        if forbit_html:
            return HttpResponse(forbit_html)

        excel_file = request.FILES.get("file")
        project = Project.objects.filter(pk=project_id).first()

        if not excel_file:
            html_message = render_message(
                request, message="Vui lý nhập file Excel", message_type="red"
            )
            return HttpResponse(html_message)
        if not project:
            html_message = render_message(
                request, message="Dự án này không tồn tại", message_type="red"
            )
            return HttpResponse(html_message)

        # Read the data from the Excel file then save as job record
        df = pd.read_excel(excel_file, header=1)

        cost_estimation_list = []
        supply_number_list = []
        category_list = []
        # Loop over each row in the Excel file
        for index, row in df.iterrows():
            if row["STT"] == "":
                continue
            cost_estimation = CostEstimation()
            cost_estimation.project = project

            # Make sure "STT", "Khối lượng", "Mã vật tư", "Ghi chú" are in the row
            if not all(
                field in row for field in ["STT", "Khối lượng", "Mã vật tư", "Ghi chú"]
            ):
                return render_modal(
                    project_id,
                    message="File excel không đúng mẫu, vui lòng kiểm tra lại",
                    message_type="red",
                )

            # check quantity
            try:
                int(row["Khối lượng"])
                cost_estimation.quantity = row["Khối lượng"]
            except:
                return render_modal(
                    project_id,
                    message="Vui lòng kiểm tra lại cột khối lượng, tất cả phải là số",
                    message_type="red",
                )

            # check note
            cost_estimation.note = row["Ghi chú"] if row["Ghi chú"] else ""
            # check if the note is nan => set it empty
            if pd.isna(cost_estimation.note):
                cost_estimation.note = ""

            # because there is a "#" at the beginning of the string to make sure the string is not number
            # so we need to remove it, also check if supply number is duplicate
            supply_number = row["Mã vật tư"][1:]
            if supply_number in supply_number_list:
                return render_modal(
                    project_id,
                    message='Mã vật tư "' + str(supply_number) + '" bị trùng lập',
                    message_type="red",
                )
            supply_number_list.append(supply_number)

            # check base_supply
            base_supply = BaseSupply.objects.filter(supply_number=supply_number).first()
            print(base_supply)
            if base_supply:
                cost_estimation.base_supply = base_supply
            else:
                return render_modal(
                    project_id,
                    message="Không tìm thấy mã vật tư: " + str(supply_number),
                    message_type="red",
                )

            # if every field is valid, append to the list
            cost_estimation_list.append(cost_estimation)

        # delete old records, save new records
        old_cost_estimations = CostEstimation.objects.filter(project=project)
        for old_cost_estimation in old_cost_estimations:
            if old_cost_estimation.get_ordered_quantity() != 0:
                # check if the old cost estimation is not in the new list, by checking if the supply_base is in the list
                if old_cost_estimation.base_supply not in [
                    cost_estimation.base_supply
                    for cost_estimation in cost_estimation_list
                ]:
                    # apeend old cost estimation to the list
                    old_cost_estimation.note = "Không có trong bảng dự toán được tải lên nhưng được giữ lại vì đã phát sinh khối lượng được đặt."
                    # make a copy to save to the list and delete from the databaes
                    cost_estimation_list.append(
                        CostEstimation(
                            project=old_cost_estimation.project,
                            base_supply=old_cost_estimation.base_supply,
                            quantity=old_cost_estimation.quantity,
                            note=old_cost_estimation.note,
                            # Copy other fields as needed
                        )
                    )
        # Delete old records after appending the necessary records
        old_cost_estimations.delete()

        # Add Vật tư phụ/ Biện pháp thi công to the list of cost_estimation_list
        # Create a BaseSupply instance for "Vật tư phụ/ Biện pháp thi công"
        auxiliary_supplies = BaseSupply.objects.filter(material_type="Vật tư phụ/ Biện pháp thi công")
        for auxiliary_supply in auxiliary_supplies:
            cost_estimation_list.append(
                CostEstimation(
                    project=project,
                    base_supply=auxiliary_supply,
                    quantity=999999,
                    note="Tự động thêm vào dự toán",
                )
            )
        # Save the records
        for cost_estimation in cost_estimation_list:
            cost_estimation.save()

        return render_modal(
            project_id, message='Cập nhật thành công. \n\n Lưu ý: "Tất cả Vật tư phụ/ Biện pháp thi công được tự động thêm vào dự toán với số lượng 999.999" ', message_type="green"
        )


def form_sub_job_cost_estimation_table(request, project_id):
    def render_modal(project_id, message=None, message_type="green"):
        # Get fields to be displayed by using record meta
        # If there is get_display_fields method, use that method
        fields = []
        headers = []
        if hasattr(SubJobEstimation, "get_display_fields"):
            for field in SubJobEstimation.get_display_fields():
                fields.append(field)
                headers.append(SubJobEstimation._meta.get_field(field).verbose_name)

        # Get the list of repair parts
        records = SubJobEstimation.objects.filter(project=project_id)
        context = {
            "records": records,
            "fields": fields,
            "headers": headers,
            "project_id": project_id,
            "message": message,
            "message_type": message_type,
        }

        return render(
            request, "components/modal_sub_job_estimation_table.html", context
        )

    if request.method == "GET":
        forbit_html = decide_permission(request, "read", {"model": "SubJobEstimation"})
        if forbit_html:
            return HttpResponse(forbit_html)
        else:
            return render_modal(project_id)

    if request.method == "POST":
        forbit_html = decide_permission(
            request, "create", {"model": "SubJobEstimation"}
        )
        if forbit_html:
            return HttpResponse(forbit_html)

        excel_file = request.FILES.get("file")
        project = Project.objects.filter(pk=project_id).first()

        if not excel_file:
            html_message = render_message(
                request, message="Vui lý nhập file Excel", message_type="red"
            )
            return HttpResponse(html_message)
        if not project:
            html_message = render_message(
                request, message="Dự án này không tồn tại", message_type="red"
            )
            return HttpResponse(html_message)

        # Read the data from the Excel file then save as job record
        df = pd.read_excel(excel_file, header=1)

        cost_estimation_list = []
        job_number_list = []
        category_list = []
        # Loop over each row in the Excel file
        for index, row in df.iterrows():
            if row["STT"] == "":
                continue
            cost_estimation = SubJobEstimation()
            cost_estimation.project = project

            # Make sure "STT", "Khối lượng", "Mã vật tư", "Ghi chú" are in the row
            if not all(
                field in row
                for field in ["STT", "Khối lượng", "Mã công việc", "Ghi chú"]
            ):
                return render_modal(
                    project_id,
                    message="File excel không đúng mẫu, vui lòng kiểm tra lại",
                    message_type="red",
                )

            # check quantity
            try:
                int(row["Khối lượng"])
                cost_estimation.quantity = row["Khối lượng"]
            except:
                return render_modal(
                    project_id,
                    message="Vui lòng kiểm tra lại cột khối lượng, tất cả phải là số",
                    message_type="red",
                )

            # check note
            cost_estimation.note = row["Ghi chú"] if row["Ghi chú"] else ""
            # check if the note is nan => set it empty
            if pd.isna(cost_estimation.note):
                cost_estimation.note = ""

            # because there is a "#" at the beginning of the string to make sure the string is not number
            # so we need to remove it, also check if supply number is duplicate
            job_number = row["Mã công việc"][1:]
            if job_number in job_number_list:
                return render_modal(
                    project_id,
                    message='Mã công việc "' + str(job_number) + '" bị trùng lập',
                    message_type="red",
                )
            job_number_list.append(job_number)

            # check base_sub_job
            base_sub_job = BaseSubJob.objects.filter(job_number=job_number).first()
            print(base_sub_job)
            if base_sub_job:
                cost_estimation.base_sub_job = base_sub_job
            else:
                return render_modal(
                    project_id,
                    message="Không tìm thấy công việc: " + str(job_number),
                    message_type="red",
                )

            # if every field is valid, append to the list
            cost_estimation_list.append(cost_estimation)

        # delete old records, save new records
        old_cost_estimations = SubJobEstimation.objects.filter(project=project)
        for old_cost_estimation in old_cost_estimations:
            if old_cost_estimation.get_ordered_quantity() != 0:
                # check if the old cost estimation is not in the new list, by checking if the supply_base is in the list
                if old_cost_estimation.base_sub_job not in [
                    cost_estimation.base_sub_job
                    for cost_estimation in cost_estimation_list
                ]:
                    # apeend old cost estimation to the list
                    old_cost_estimation.note = "Không có trong bảng dự toán được tải lên nhưng được giữ lại vì đã phát sinh khối lượng được đặt."
                    # make a copy to save to the list and delete from the databaes
                    cost_estimation_list.append(
                        SubJobEstimation(
                            project=old_cost_estimation.project,
                            base_supply=old_cost_estimation.base_sub_job,
                            quantity=old_cost_estimation.quantity,
                            note=old_cost_estimation.note,
                            # Copy other fields as needed
                        )
                    )
        # Delete old records after appending the necessary records
        old_cost_estimations.delete()

        # Save the records
        for cost_estimation in cost_estimation_list:
            cost_estimation.save()

        return render_modal(
            project_id, message="Cập nhật thành công", message_type="green"
        )


def form_maintenance_images(request, maintenance_id):
    # If Get
    if request.method == "GET":
        vehicle_maintenance = VehicleMaintenance.objects.get(pk=maintenance_id)
        # Get the list of images for the maintenance record
        maintenance_images = MaintenanceImage.objects.filter(
            vehicle_maintenance=vehicle_maintenance
        )

        context = {
            "maintenance_id": maintenance_id,
            "maintenance_images": maintenance_images,
        }
        return render(request, "components/modal_gallery.html", context)
    # If Post
    elif request.method == "POST":
        images = []
        try:
            # Save images to the database
            for image in request.FILES.getlist("images[]"):
                maintenance_image = MaintenanceImage.objects.create(
                    vehicle_maintenance_id=maintenance_id, image=image
                )
                images.append(
                    {"id": maintenance_image.id, "url": maintenance_image.image.url}
                )
            return JsonResponse({"success": True, "images": images})
        except:
            return JsonResponse({"success": False})


def form_maintenance_payment_request(request):
    # If Get
    if request.method == "GET":
        # Get all PaymentRecord
        payments = PaymentRecord.objects.all()
        # put them into group of payments with the keys are vehicle_maintenance and provider
        groups = {}
        for payment in payments:
            key = (payment.vehicle_maintenance_id, payment.provider_id)
            if key in groups:
                groups[key].append(payment)
            else:
                groups[key] = [payment]

        context = {"groups": groups}
        return render(request, "components/maintenance_payment_request.html", context)
    # If Post
    elif request.method == "POST":
        return JsonResponse({"success": True})


# PAGES ==============================================================
@login_required
def page_home(request, sub_page=None):

    if sub_page is None:
        return redirect("page_home", sub_page="Announcement")

    display_name_dict = {
        "Announcement": "Thông báo",
        "User": User.get_vietnamese_name(),
        "Permission": Permission.get_vietnamese_name(),
        "ProjectUser": ProjectUser.get_vietnamese_name(),
    }

    context = {
        "sub_page": sub_page,
        "display_name_dict": display_name_dict,
        "model": sub_page,
        "current_url": request.path,
        "header_title": display_name_dict.get(sub_page, "Trang chủ"),
    }
    return render(request, "pages/page_home.html", context)


@login_required
def page_general_data(request, sub_page=None):
    if sub_page == None:
        return redirect("page_general_data", sub_page="VehicleType")

    display_name_dict = {
        "VehicleType": "DL loại xe",
        "VehicleRevenueInputs": "DL tính DT theo loại xe",
        "VehicleDetail": "DL xe chi tiết",
        "StaffData": "DL nhân viên",
        "DriverSalaryInputs": "DL mức lương tài xế",
        "DumbTruckPayRate": "DL tính lương tài xế xe ben",
        "DumbTruckRevenueData": "DL tính DT xe ben",
        "Location": "DL địa điểm",
        "NormalWorkingTime": "Thời gian làm việc",
        "Holiday": "Ngày lễ",
    }
    context = {
        "sub_page": sub_page,
        "model": sub_page,
        "display_name_dict": display_name_dict,
        "current_url": request.path,
        "header_title": display_name_dict.get(sub_page, "Dữ liệu chung"),
    }
    return render(request, "pages/page_general_data.html", context)


@login_required
def page_transport_department(request, sub_page=None):
    if sub_page == None:
        return redirect("page_transport_department", sub_page="LiquidUnitPrice")

    display_name_dict = {
        "LiquidUnitPrice": "Bảng đơn giá nhiên liệu/nhớt",
        "FillingRecord": "LS đổ nhiên liệu/nhớt",
        "PartProvider": "Nhà cung cấp phụ tùng",
        "RepairPart": "Danh mục phụ tùng",
        "PaymentRecord": "LS thanh toán",
        "VehicleMaintenance": "Phiếu sửa chữa",
        "VehicleDepreciation": "Khấu hao",
        "VehicleBankInterest": "Lãi ngân hàng",
        "VehicleOperationRecord": "DL HĐ xe công trình / ngày",
        "ConstructionDriverSalary": "Bảng lương",
        "ConstructionReportPL": "Bảng BC P&L xe cơ giới",
    }

    params = request.GET.copy()
    if "start_date" not in params:
        start_date = get_valid_date("")
    else:
        start_date = params["start_date"]
    if "end_date" not in params:
        end_date = get_valid_date("")
    else:
        end_date = params["end_date"]
    if "check_month" not in params:
        check_month = get_valid_month("")
    else:
        check_month = params["check_month"]

    context = {
        "sub_page": sub_page,
        "model": sub_page,
        "display_name_dict": display_name_dict,
        "current_url": request.path,
        "start_date": start_date,
        "end_date": end_date,
        "check_month": check_month,
        "header_title": display_name_dict.get(sub_page, "Phòng vận tải"),
    }
    return render(request, "pages/page_transport_department.html", context)


@login_required
def page_projects(request, sub_page=None):
    if sub_page == None:
        return redirect("page_projects", sub_page="Project")

    display_name_dict = {
        "Project": "Dự án",
        "SupplyProvider": "Nhà cung cấp vật tư",
        "SupplyBrand": "Thương hiệu vật tư",
        "BaseSupply": "Vật tư",
        "DetailSupply": "Vật tư chi tiết",
        "SupplyPaymentRecord": "LS thanh toán vật tư",
        "SubContractor": "Tổ đội/ nhà thầu phụ",
        "BaseSubJob": "Công việc của tổ đội/ nhà thầu phụ",
        "DetailSubJob": "Công việc chi tiết của tổ đội/ nhà thầu phụ",
        "SubJobPaymentRecord": "LS thanh toán công việc của tổ đội/ nhà thầu phụ",
    }
    context = {
        "sub_page": sub_page,
        "model": sub_page,
        "display_name_dict": display_name_dict,
        "current_url": request.path,
        "header_title": display_name_dict.get(sub_page, "Dự án"),
    }
    return render(request, "pages/page_projects.html", context)


@login_required
def page_each_project(request, pk):
    check_date = request.GET.get("check_date")
    project_id = get_valid_id(pk)
    project = get_object_or_404(Project, pk=project_id)
    # Should check if the project is belong to the user
    context = {"project_id": project_id, "check_date": check_date, "project": project}
    return render(request, "pages/page_each_project.html", context)


def clean(request):
    result = ""
    # Find MaintenanceImage records with null vehicle_maintenance
    orphaned_images = MaintenanceImage.objects.filter(vehicle_maintenance__isnull=True)
    # Delete the orphaned records
    deleted_count = orphaned_images.count()
    orphaned_images.delete()
    result = f"Deleted {deleted_count} MaintenanceImage records with null foreign keys."

    orphan_records = VehicleMaintenanceRepairPart.objects.filter(
        vehicle_maintenance__isnull=True
    )
    # Delete the orphaned records
    deleted_count = orphan_records.count()
    orphan_records.delete()
    result += f"\nDeleted {deleted_count} VehicleMaintenanceRepairPart records with null foreign keys"

    # calculate PartProvider
    records = PartProvider.objects.all()
    for record in records:
        record.save()

    # do the same for supply
    orphan_records = SupplyPaymentRecord.objects.filter(supply_order__isnull=True)
    # Delete the orphaned records
    deleted_count = orphan_records.count()
    orphan_records.delete()
    result += (
        f"\nDeleted {deleted_count} SupplyPaymentRecord records with null foreign keys"
    )

    # supply order supply
    orphan_records = SupplyOrderSupply.objects.filter(supply_order__isnull=True)
    # Delete the orphaned records
    deleted_count = orphan_records.count()
    orphan_records.delete()
    result += (
        f"\nDeleted {deleted_count} SupplyOrderSupply records with null foreign keys"
    )

    return HttpResponse(result)


def gps(request):
    return render(request, "pages/gps.html")
